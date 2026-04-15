import os
import json
from openpyxl import Workbook, load_workbook


class OrderManager:
    def __init__(self):
        self.history = []
        self.load_history()

    def load_history(self):
        if os.path.exists("history.json"):
            with open("history.json", "r", encoding="utf-8") as f:
                self.history = json.load(f)

    def save_history(self):
        with open("history.json", "w", encoding="utf-8") as f:
            json.dump(self.history, f, ensure_ascii=False, indent=4)

    def save_to_excel(self, order):
        file_path = "order_history.xlsx"

        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.append(["Time", "Item", "Price"])
        else:
            wb = load_workbook(file_path)
            ws = wb.active

        for item in order["items"]:
            name, price = item.split("-")
            price = int(price.replace("đ", "").strip())
            ws.append([order["time"], name.strip(), price])

        wb.save(file_path)